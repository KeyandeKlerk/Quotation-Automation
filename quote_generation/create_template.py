# File to create the basic template of the quote

import process_spreadsheet
import random
import os

from openpyxl import Workbook
from openpyxl.styles import Border, Side
from datetime import datetime


# Function that initiates the required variables


def init():
    global workbook, sheet

    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 5


# Function to load json


def insert_data_from_list(json_data):
    counter = 1

    for products in json_data:
        append_row([counter, products[1], "", products[2], products[3]])
        row_num = sheet.max_row
        merge_cells("B" + str(row_num) + ":C" + str(row_num))
        sheet["F" + str(row_num)] = "=D" + str(row_num) + "*E" + str(row_num)

        counter += 1


# Function to merge cells in range


def merge_cells(range):

    sheet.merge_cells(range)


# Function to add row to table


def append_row(value):
    # Value must be a list
    sheet.append(value)


# Function to add border to range of cells


def set_border(cell_range):
    thin = Side(border_style="medium", color="000000")
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


# Function that creates template head


def template_head(client_name, requested_by, area_on_site, contact_info, work):

    init()

    current_date = datetime.now()
    current_date = current_date.strftime("%Y-%m-%d")

    append_row(["BILL OF QUANTITIES"])
    append_row(["CLIENT: " + client_name, "", "", "REQUESTED BY: " + requested_by])
    append_row(
        ["AREA on SITE: " + area_on_site, "", "", "Contact Info: " + contact_info]
    )
    append_row(["WORK: " + work])
    append_row(["TECHNICIAN: JP", "", "", "", "Date: " + current_date])
    append_row(["ITEM", "DESCRIPTION", "", "QTY", "UNIT", "TOTAL"])
    append_row([""])

    merge_cells("A1:G1")
    merge_cells("A2:C2")
    merge_cells("A3:C3")
    merge_cells("D2:G2")
    merge_cells("D3:G3")
    merge_cells("A4:G4")
    merge_cells("A5:D5")
    merge_cells("E5:G5")
    merge_cells("B6:C6")
    merge_cells("A7:G7")

    set_border("A1:G6")


# Function that creates category A


def create_category_a(json_data, markup):
    append_row(["A"])

    # Loop to populate category A
    insert_data_from_list(json_data)
    current_rows_used = sheet.max_row
    append_row([""])

    append_row(["", "Cost"])
    append_row(["", "Markup", "", markup])
    append_row(["", "Sub Total Excl V.A.T"])

    sheet["G" + str(current_rows_used + 2)] = "=SUM(F9:F" + str(current_rows_used) + ")"
    sheet["G" + str(current_rows_used + 3)] = (
        "=G" + str(current_rows_used + 2) + "* D" + str(current_rows_used + 3)
    )
    sheet["G" + str(current_rows_used + 4)] = (
        "=SUM(G" + str(current_rows_used + 2) + ":G" + str(current_rows_used + 3) + ")"
    )

    global categoryA_total, categoryA_markup, categoryA_subtotal_excl

    categoryA_total = "G" + str(current_rows_used + 2)
    categoryA_markup = "G" + str(current_rows_used + 3)
    categoryA_subtotal_excl = "G" + str(current_rows_used + 4)

    merge_cells("B8:G8")
    merge_cells("B" + str(current_rows_used + 1) + ":C" + str(current_rows_used + 1))
    merge_cells("B" + str(current_rows_used + 2) + ":C" + str(current_rows_used + 2))
    merge_cells("B" + str(current_rows_used + 3) + ":C" + str(current_rows_used + 3))
    merge_cells("B" + str(current_rows_used + 4) + ":C" + str(current_rows_used + 4))


# Function that creates category B


def create_category_b(json_data, markup):

    append_row(["B"])
    rows_before_population = sheet.max_row
    # Loop to populate category A
    insert_data_from_list(json_data)
    current_rows_used = sheet.max_row
    append_row([""])

    append_row(["", "Cost"])
    append_row(["", "Markup", "", markup])
    append_row(["", "Sub Total Excl V.A.T"])

    sheet["G" + str(current_rows_used + 2)] = (
        "=SUM(F" + str(rows_before_population + 1) + ":F" + str(current_rows_used) + ")"
    )
    sheet["G" + str(current_rows_used + 3)] = (
        "=G" + str(current_rows_used + 2) + "* D" + str(current_rows_used + 3)
    )
    sheet["G" + str(current_rows_used + 4)] = (
        "=SUM(G" + str(current_rows_used + 2) + ":G" + str(current_rows_used + 3) + ")"
    )

    global categoryB_total, categoryB_markup, categoryB_subtotal_excl

    categoryB_total = "G" + str(current_rows_used + 2)
    categoryB_markup = "G" + str(current_rows_used + 3)
    categoryB_subtotal_excl = "G" + str(current_rows_used + 4)

    merge_cells("B" + str(rows_before_population) + ":G" + str(rows_before_population))
    merge_cells("B" + str(current_rows_used + 1) + ":C" + str(current_rows_used + 1))
    merge_cells("B" + str(current_rows_used + 2) + ":C" + str(current_rows_used + 2))
    merge_cells("B" + str(current_rows_used + 3) + ":C" + str(current_rows_used + 3))
    merge_cells("B" + str(current_rows_used + 4) + ":C" + str(current_rows_used + 4))


# Function that creates category C


def create_category_c(json_data, markup):

    append_row(["C"])
    rows_before_population = sheet.max_row
    # Loop to populate category A
    insert_data_from_list(json_data)
    current_rows_used = sheet.max_row

    append_row([""])
    append_row(["", "Cost"])
    append_row(["", "Markup", "", markup])
    append_row(["", "Sub Total Excl V.A.T"])
    append_row([""])

    sheet["G" + str(current_rows_used + 2)] = (
        "=SUM(F" + str(rows_before_population + 1) + ":F" + str(current_rows_used) + ")"
    )
    sheet["G" + str(current_rows_used + 3)] = (
        "=G" + str(current_rows_used + 2) + "* D" + str(current_rows_used + 3)
    )
    sheet["G" + str(current_rows_used + 4)] = (
        "=SUM(G" + str(current_rows_used + 2) + ":G" + str(current_rows_used + 3) + ")"
    )

    global categoryC_total, categoryC_markup, categoryC_subtotal_excl

    categoryC_total = "G" + str(current_rows_used + 2)
    categoryC_markup = "G" + str(current_rows_used + 3)
    categoryC_subtotal_excl = "G" + str(current_rows_used + 4)

    merge_cells("B" + str(rows_before_population) + ":G" + str(rows_before_population))
    merge_cells("B" + str(current_rows_used + 1) + ":C" + str(current_rows_used + 1))
    merge_cells("B" + str(current_rows_used + 2) + ":C" + str(current_rows_used + 2))
    merge_cells("B" + str(current_rows_used + 3) + ":C" + str(current_rows_used + 3))
    merge_cells("B" + str(current_rows_used + 4) + ":C" + str(current_rows_used + 4))
    merge_cells("A" + str(current_rows_used + 5) + ":G" + str(current_rows_used + 5))


# Function that creates final totals and notes


def final_totals(quote_name, notes=""):

    curr_dir = os.getcwd()
    filename = curr_dir + "\\generated quotes\\" + quote_name + ".xlsx"
    current_rows_used = sheet.max_row

    append_row(["", "Total Cost Excl. VAT"])
    append_row(["", "Total Markup Excl. VAT"])
    append_row(["", "Total Excl. VAT"])
    append_row(["", "VAT", "", "15%"])
    append_row(["", "Total Incl. VAT"])
    append_row([""])
    append_row(["NOTES: " + notes])

    sheet["G" + str(current_rows_used + 1)] = (
        "=SUM(" + categoryA_total + "," + categoryB_total + "," + categoryC_total + ")"
    )
    sheet["G" + str(current_rows_used + 2)] = (
        "=SUM("
        + categoryA_markup
        + ","
        + categoryB_markup
        + ","
        + categoryC_markup
        + ")"
    )
    sheet["G" + str(current_rows_used + 3)] = (
        "=SUM(G" + str(current_rows_used + 1) + ",G" + str(current_rows_used + 2) + ")"
    )
    sheet["G" + str(current_rows_used + 4)] = (
        "=SUM(G" + str(current_rows_used + 3) + "* D" + str(current_rows_used + 4) + ")"
    )
    sheet["G" + str(current_rows_used + 5)] = (
        "=SUM(G" + str(current_rows_used + 3) + ":G" + str(current_rows_used + 4) + ")"
    )

    merge_cells("B" + str(current_rows_used + 1) + ":C" + str(current_rows_used + 1))
    merge_cells("B" + str(current_rows_used + 2) + ":C" + str(current_rows_used + 2))
    merge_cells("B" + str(current_rows_used + 3) + ":C" + str(current_rows_used + 3))
    merge_cells("B" + str(current_rows_used + 4) + ":C" + str(current_rows_used + 4))
    merge_cells("B" + str(current_rows_used + 5) + ":C" + str(current_rows_used + 5))
    merge_cells("A" + str(current_rows_used + 6) + ":G" + str(current_rows_used + 6))
    merge_cells("A" + str(current_rows_used + 7) + ":G" + str(current_rows_used + 13))

    workbook.save(filename)
    workbook.close()
