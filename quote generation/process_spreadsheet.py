# File to process the excel spreadsheet and save each sheet as a separate json file

import json
import os

import openpyxl

# Starting point of file


def main():

    global priceList1, curr_dir
    curr_dir = os.getcwd()
    priceList1 = openpyxl.load_workbook(
        f'{curr_dir}\\spreadsheets\\Template.xlsx', data_only=True)

    # Function to process each individual sheet

    process_category_a(priceList1)
    process_category_b(priceList1)
    process_category_c(priceList1)

# Function to process Conventional Products Sheet


def process_category_a(priceList1):

    activeSheet = priceList1['CategoryA']
    cells = []
    rows = []

    for row in activeSheet.iter_rows(min_row=2, min_col=1):

        for cell in row:

            cells.append(str(cell.value))

        rows.append(cells)
        cells = []

    save_to_file(rows, 'categoryA.json')

# Function to process Specialist Products Sheet


def process_category_b(priceList1):

    activeSheet = priceList1['CategoryB']
    cells = []
    rows = []

    for row in activeSheet.iter_rows(min_row=2, min_col=1):

        for cell in row:

            cells.append(str(cell.value))

        rows.append(cells)
        cells = []

    save_to_file(rows, 'categoryB.json')

# Function to process Intelligent Equipment Sheet


def process_category_c(priceList1):

    activeSheet = priceList1['CategoryC']
    cells = []
    rows = []

    for row in activeSheet.iter_rows(min_row=2, min_col=1):

        for cell in row:

            cells.append(str(cell.value))

        rows.append(cells)
        cells = []

    save_to_file(rows, 'categoryC.json')

# Function to save json data to file


def save_to_file(json_data, filename):

    json_filename = curr_dir + '\\json\\' + filename

    with open(json_filename, 'w') as outfile:

        json.dump(json_data, outfile)
